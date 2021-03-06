import matplotlib.pyplot as plt
import numpy
from numpy import genfromtxt
import csv
import pandas as pd
from operator import itemgetter
from datetime import*
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import openpyxl
from win32com import client
#######################################################################
def nearest(items, pivot):
    return min(items, key=lambda x: abs(x - pivot))
#######################################################################
def add_one_month(t):
    """Return a `datetime.date` or `datetime.datetime` (as given) that is
    one month earlier.

    Note that the resultant day of the month might change if the following
    month has fewer days:

        >>> add_one_month(datetime.date(2010, 1, 31))
        datetime.date(2010, 2, 28)
    """
    import datetime
    one_day = datetime.timedelta(days=1)
    one_month_later = t + one_day
    while one_month_later.month == t.month:  # advance to start of next month
        one_month_later += one_day
    target_month = one_month_later.month
    while one_month_later.day < t.day:  # advance to appropriate day
        one_month_later += one_day
        if one_month_later.month != target_month:  # gone too far
            one_month_later -= one_day
            break
    return one_month_later
#######################################################################
def subtract_one_month(t):
    """Return a `datetime.date` or `datetime.datetime` (as given) that is
    one month later.

    Note that the resultant day of the month might change if the following
    month has fewer days:

        >>> subtract_one_month(datetime.date(2010, 3, 31))
        datetime.date(2010, 2, 28)
    """
    import datetime
    one_day = datetime.timedelta(days=1)
    one_month_earlier = t - one_day
    while one_month_earlier.month == t.month or one_month_earlier.day > t.day:
        one_month_earlier -= one_day
    return one_month_earlier
#######################################################################
values=[]
dates=[]
combine=[]
with open('hyatt.csv', 'r') as csvFile:
    reader = csv.reader(csvFile)
    for row in reader:
        values.append(row[1])
        dates.append(row[1])
        combine.append(row)
csvFile.close()
#print(values)
#print(dates)
#print(combine)
print('The Number of Values are: ',len(values))
print('The Number of Dates are:',len(dates))
combine = sorted(combine, key=itemgetter(0))
"""for i in combine:
    print(i)"""
for i in combine:
    m2=i[0]
    m2=datetime.strptime(m2,'%d/%m/%y %I:%M %p')
    i[0]=m2
combine = sorted(combine, key=itemgetter(0))
"""for i in combine:
    print(i)"""
ref_min=combine[0][0].date()
min_time=datetime.strptime('0000','%H%M').time()
ref_min=datetime.combine(ref_min, min_time)
print(type(ref_min))
ref_max=combine[-1][0].date()
max_time=datetime.strptime('2359','%H%M').time()
ref_max=datetime.combine(ref_max, max_time)
print(ref_max)
ref_max = add_one_month(ref_max)
dates=[]
for i in combine:
    dates.append(i[0])
i=ref_min
indices=[]
while i<ref_max:
    k=nearest(dates,i)
    print('The corresponding Lowest time related to this reading is: ',k)
    index=dates.index(k)
    print(index)
    indices.append(index)
    i = add_one_month(i)
    print(i)
print('The Number of Indices are: ',len(indices))
k=ref_min.date()
consump=[]
lower=[]
upper=[]
for i in range(len(indices)-1):
    r_min=float(values[indices[i]])
    lower.append(r_min)
    r_up=float(values[indices[i+1]])
    upper.append(r_up)
    consumption=r_up-r_min
    consump.append(consumption)
    print('The Consumption on ',k.strftime('%d-%m-%Y'),' is : ',consumption)
    k = add_one_month(k)
rate=float(input('Enter the Rate per KWH consumed for Cost Calculation: '))
k=ref_min.date()
cost=[]
for i in consump:
    r=float(i)*rate
    print('The Cost of Electricity for ',k.strftime('%d-%m-%Y'),' is :',r)
    cost.append(r)
    k = add_one_month(k)
print('\n====================Final Output====================\n')
k=ref_min.date()
date_list=[]
write=[]
write2=[]
cust=input('Please Enter The Customer Name:  ')
row=['Customer Name: ',cust]
write.append(row)
row=['Address Line 1: ',"3, National Hwy 9, Premnagar, "]
write.append(row)
row=['Address Line 2: ',"Ashok Nagar, Pune, Maharashtra 411016"]
write.append(row)
row=['']
write.append(row)
row=['Electricity Bill Invoice']
write.append(row)
row=['From: ',ref_min.date()]
write.append(row)
row=['To: ',subtract_one_month(ref_max.date())+timedelta(days=1)]
write.append(row)
row=['']
write.append(row)
row=['Reading Date','Previous Reading','Present Reading','Consumption','Cost']
write.append(row)
for i in range(len(indices)-1):
    row=[]
    row2=[]
    print('--------------------------------------')
    print('Date:\t\t',k)
    row.append(k)
    row2.append(k)
    date_list.append(k)
    k = add_one_month(k)
    print('Lower Reading:\t',lower[i])
    row.append(lower[i])
    row2.append(lower[i])
    print('Upper Reading:\t',upper[i])
    row.append(upper[i])
    row2.append(upper[i])
    print('Consumption:\t',consump[i])
    row.append(consump[i])
    row2.append(consump[i])
    print('Cost:\t\t',cost[i])
    row.append(cost[i])
    row2.append(cost[i])
    write.append(row)
    write2.append(row2)
plt.plot(date_list,consump)
plt.show()
plt.savefig('graph.png')
row=['Total Consumption: ', sum(consump)]
write.append(row)
row=['Cost Per Unit: ',rate]
write.append(row)
row=['Total Bill Ammount: ',sum(cost)]
write.append(row)
with open('output.csv', 'w') as csvFile:
    for row in write:
        writer = csv.writer(csvFile,lineterminator='\n')
        writer.writerow(row)
csvFile.close()
print('CSV FILE Generated as Output.csv')
###########################################################################
wb=load_workbook('Book1.xlsx')
ws1=wb.get_sheet_by_name('Sheet1')
# shs is list
ws1['B2']=cust
ws1['B3']='3, National Hwy 9, Premnagar, '
ws1['B4']='Ashok Nagar, Pune, Maharashtra 411016'
ws1['B6']=ref_min.date()
ws1['E6']=subtract_one_month(ref_max.date())+timedelta(days=1)
row=9
column=1
for r in write2:
    column=1
    for i in r:
        ws1.cell(row,column).value=i
        column+=1
    row+=1
"""
row+=1
column=1
ws1.cell(row,column).value='Total Consumption: '
ws1.cell(row,column).font=Font(bold=True)
column+=1
ws1.cell(row,column).value=sum(consump)
column-=1
row+=1
ws1.cell(row,column).value='Total Cost: '
ws1.cell(row,column).font=Font(bold=True)
column+=1
ws1.cell(row,column).value=sum(cost)
column-=1"""
thick_border_right=Border(right=Side(style='thick'))
ws1['E2'].border=thick_border_right
ws1['E3'].border=thick_border_right
ws1['E4'].border=thick_border_right
thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
ws1['A15']='Total Consumption'
ws1['A15'].font=Font(bold=True)
ws1['A15'].border=thick_border
ws1['B15'].border=thick_border
ws1['C15'].border=thick_border
ws1['D15'].border=thick_border
ws1['A16']='Total Cost'
ws1['A16'].font=Font(bold=True)
ws1['A16'].border=thick_border
ws1['B16'].border=thick_border
ws1['C16'].border=thick_border
ws1['D16'].border=thick_border
ws1['E15']=sum(consump)
ws1['E16']=sum(cost)
img = openpyxl.drawing.image.Image('logo.jpg')
img.anchor='A1'
ws1.add_image(img)
wb.save('Book1.xlsx')
#############################################################################
xlApp = client.Dispatch("Excel.Application")
books = xlApp.Workbooks.Open('E:\Internship\Siemens\EMAPP\Book1.xlsx')
ws = books.Worksheets[0]
ws.Visible = 1
ws.ExportAsFixedFormat(0, 'E:\Internship\Siemens\EMAPP\trial.pdf')
